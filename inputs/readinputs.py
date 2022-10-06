import openpyxl
import ast
import json
import pandas as pd
import os
import datetime

__location__ = os.path.realpath(
    os.path.join(os.getcwd(), os.path.dirname(__file__)))

def input_general(ws1,inputpath):
    
    # Worksheet 1: all inputs except for electricity prices
    # Each column is a different case for which inputs are created
    
    # Read ws 1 and store all values in dict1
    
    dict1 = {}
    
    for col in range(2,ws1.max_column + 1):
        
        keycol = ws1.cell(1,col).value
        dict1[keycol] = {}
    
        key1 = 'init'
    
        for row in range(2,ws1.max_row + 1):
            
            key2 = ws1.cell(row, 1).value
            value = ws1.cell(row, col).value
            
            if key2 in ['general','economics','pv','battery','inv']:
                key1 = key2
                dict1[keycol][key1] = {}
                
            else:
            
                if key2 in ['columns','TechsShift','location']:
                    value = ast.literal_eval(value)
                
                if key2 in ['WetAppManualShifting','PresenceOfPV','PresenceOfBattery','AutomaticSizing', 'TMY']:
                    value = bool(value)
                    
                dict1[keycol][key1][key2] = value

    # Save read values in corresponding json files

    cases = {}
    for i in dict1.keys():
        cases[i]=dict1[i]['general'].copy()
    filename = inputpath + 'cases.json'
    with open(filename, 'w',encoding='utf-8') as f:
        json.dump(cases, f,ensure_ascii=False, indent=4)
        
    economics = {}
    for i in dict1.keys():
        economics[i]=dict1[i]['economics'].copy()
    filename = inputpath + 'econ_param.json'
    with open(filename, 'w',encoding='utf-8') as f:
        json.dump(economics, f,ensure_ascii=False, indent=4)
        
    pvbatt = {}
    for i in dict1.keys():
        pvbatt[i] = {}
        pvbatt[i]['pv']=dict1[i]['pv'].copy()
        pvbatt[i]['battery']= dict1[i]['battery'].copy()
        pvbatt[i]['inv']= dict1[i]['inv'].copy()
    filename = inputpath + 'pvbatt_param.json'
    with open(filename, 'w',encoding='utf-8') as f:
        json.dump(pvbatt, f,ensure_ascii=False, indent=4)
        
    return dict1



def input_elprices(ws2,inputpath):
    
    # Worksheet 2: electricity prices
    # Electricity prices are the same for all cases defined in ws1
    # Users can change only the type of tariff (single, double, multi price) per each case
    
    # Read ws 2 and store all values in dict2            
    
    dict2 = {}
    col = 2
    for row in range(1,ws2.max_row + 1):
        
        key = ws2.cell(row, 1).value
        value = ws2.cell(row, col).value
        dict2[key] = value
    
    # Create yearly arrays of prices and save as csv
    
    index = pd.date_range(start='2015-01-01',end='2015-12-31 23:59:00',freq='15T')
    
    times = {}
    prices_en = {}
    prices_grid = {}
    
    names = ['single','double','multi']
    
    times['single']       = [['0:00','23:59']]
    prices_en['single']   = [dict2['single_en']]
    prices_grid['single'] = [dict2['single_grid']]
    
    times['double']       = [['0:00','6:00'],['6:00','11:00'],['11:00','17:00'],['17:00','22:00'],['22:00','23:59']]
    prices_en['double']   = [dict2['double_en_low'],  dict2['double_en_high'],  dict2['double_en_low'],  dict2['double_en_high'],  dict2['double_en_low']]
    prices_grid['double'] = [dict2['double_grid_low'],dict2['double_grid_high'],dict2['double_grid_low'],dict2['double_grid_high'],dict2['double_grid_low']]
    
    times['multi']       = [['0:00','6:00'],['6:00','11:00'],['11:00','17:00'],['17:00','22:00'],['22:00','23:59']]
    prices_en['multi']   = [dict2['multi_en_midlow'],  dict2['multi_en_midhigh'],  dict2['multi_en_low'],  dict2['multi_en_high'],  dict2['multi_en_midlow']]
    prices_grid['multi'] = [dict2['multi_grid_midlow'],dict2['multi_grid_midhigh'],dict2['multi_grid_low'],dict2['multi_grid_high'],dict2['multi_grid_midlow']]
    
    for i in names:
        
        df = pd.DataFrame(index=index,columns=['energy','grid','sell'],dtype='float64')
        df['sell'] = dict2['sell']
        
        for j in range(len(times[i])):
            
            df['energy'][df.between_time(times[i][j][0],times[i][j][1]).index] = prices_en[i][j]
            df['grid'][df.between_time(times[i][j][0],times[i][j][1]).index] = prices_grid[i][j]
            name = inputpath + i + '_price.csv'
            df.to_csv(name)          
     
def read_sheet(file,sheet):
    '''
    function that reads one sheet of the excel config file and outputs it in a dataframe
    '''
    raw = pd.read_excel(file,sheet_name=sheet)
    raw.rename(columns={ raw.columns[0]: "varname" }, inplace = True)
    raw = raw.loc[raw['varname'].notna(),:]
    raw.index = raw['varname']
    return raw[['Variable','Valeur','Description']]


def read_config(filename):
    '''
    Function that read the excel config file for the load-shifting library
    Parameters
    ----------
    filename : string
        path to the config file

    Returns
    -------
    dict
    '''
    config = read_sheet(filename,'main')
    ownership = read_sheet(filename,'ownership')
    
    sellprice_2D = pd.read_excel(filename,sheet_name='sellprice',index_col=0)
    gridprice_2D = pd.read_excel(filename,sheet_name='gridprice',index_col=0)
    
    idx = pd.date_range(start=datetime.datetime(year = sellprice_2D.index[0].year, month = 1, day = 1,hour=0),end = datetime.datetime(year = sellprice_2D.index[0].year, month = 12, day = 31,hour=23),freq='1h')
    
    # turn data into a column (pd.Series)
    sellprice = sellprice_2D.stack()
    sellprice.index = idx
    
    gridprice = gridprice_2D.stack()
    gridprice.index = idx

    return {'config':config,'ownership':ownership,'gridprice':gridprice,'sellprice':sellprice}    

if __name__ == '__main__':
    
    # Open workbook and select worksheets
    wb = openpyxl.load_workbook('./inputs.xlsx')
    ws1 = wb['inputs']
    ws2 = wb['elprices']
    
    inputpath =  __location__ + '/'

    input_general(ws1,inputpath)
    input_elprices(ws2,inputpath)    
    
    # load the new, all-included config file:
    filename = './config.xlsx'
    conf = read_config(filename)
    

    
























